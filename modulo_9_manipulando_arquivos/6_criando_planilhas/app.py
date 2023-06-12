"""
Sheets - São as páginas de uma planilha
Workbook - É o arquivo que contém estas páginas
"""

import openpyxl as excel
# Criar planilha
planilha_teste = excel.Workbook()
planilha_teste.create_sheet('Frutas')
planilha_teste.create_sheet('Legumes')
planilha_teste.create_sheet('Sementes')
print(planilha_teste.sheetnames)

# Cria o cabeçalho da planilha
cabecalho_frutas = ['Título', 'Localização', 'Preço']
sheet_frutas = planilha_teste.get_sheet_by_name('Frutas')
sheet_frutas.append(cabecalho_frutas)

# Salvando dados na planilha
frutas = [
    ['Uva', 'Mercado', 5],
    ['Melancia', 'Mercado', 15],
    ['Bolo', 'Mercado', 40]
]

for fruta in frutas:
    sheet_frutas.append(fruta)

    
# Salva a planilha criada
planilha_teste.save('Dados Supermercado.xlsx')