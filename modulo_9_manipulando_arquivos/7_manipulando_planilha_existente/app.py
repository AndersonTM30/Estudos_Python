import openpyxl as excel

def criar_planilha():
    workbook = excel.Workbook()
    workbook.create_sheet('Pessoas')
    cabecalho_pessoas = ['Id', 'First Name', 'Last Name', 'Gender', 'Country', 'Age']
    sheet_pessoas = workbook.get_sheet_by_name('Pessoas')
    sheet_pessoas.append(cabecalho_pessoas)

    pessoas = [
        ['1', 'Anderson', 'Ferreira', 'Masculino', 'Brasil', '31'],
        ['2', 'Carol', 'Lima', 'Feminino', 'Brasil', '20'],
        ['3', 'Jeff', 'Albert', 'Masculino', 'Estados Unidos', '38'],
        ['4', 'Lara', 'Croft', 'Feminino', 'Inglaterra', '45'],
        ['5', 'Albert', 'Eistein', 'Masculino', 'Alemanha', '100'],
    ]
    
    for pessoa in pessoas:
        sheet_pessoas.append(pessoa)

    
    workbook.save('Pessoas.xlsx')

# criar_planilha();

planilha = excel.load_workbook('Pessoas.xlsx')
print(planilha.sheetnames)
pessoas = planilha.get_sheet_by_name('Pessoas') # armazena os valores de uma página da planilha
print(pessoas['C3'].value)

pessoas['C3'].value = 'Alves' # Altera o valor e um local específico da planilha

print(pessoas['C3'].value)

pessoas.cell(row=3,column=3,value='Mark') # Forma mais legível de alterar um valor dentro da planilha

print(pessoas['C3'].value)

# Como percorrer uma planilha?
for linha in pessoas.iter_rows(min_row=2, max_row=5,min_col=2, max_col=4):
    print(linha[0].value, linha[1].value, linha[2].value)

# Imprime todos os dados de uma coluna
for linha in pessoas.iter_rows(min_col=2, max_col=2, min_row=2):
    for cell in linha:
        print(cell.value)